import openpyxl, json
path=r'C:\\Users\\user\\Desktop\\bday ni yna\\Attendance-Year-2025.xlsx'
wb=openpyxl.load_workbook(path, read_only=True)
print('sheets:', wb.sheetnames)
ws=wb[wb.sheetnames[0]]
print('first sheet title:', ws.title)
print('max_row', ws.max_row, 'max_col', ws.max_column)
header_i=None
header_row=None
for i,row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
    if row and any(isinstance(c,str) and 'TEAM OF 12 MEN' in c for c in row):
        header_i=i; header_row=row; break
print('header_i', header_i)
if header_i:
    men_start=women_start=None
    for idx,val in enumerate(header_row):
        if isinstance(val,str) and 'TEAM OF 12 MEN' in val:
            men_start=idx
        if isinstance(val,str) and 'TEAM OF 12 WOMEN' in val:
            women_start=idx
    print('men_start', men_start, 'women_start', women_start)
    leaders_row = header_i+1
    leaders = list(ws.iter_rows(min_row=leaders_row, max_row=leaders_row, values_only=True))[0]
    print('leaders row sample men', leaders[men_start:men_start+10])
    print('leaders row sample women', leaders[women_start:women_start+10])
